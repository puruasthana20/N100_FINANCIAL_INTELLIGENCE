from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from src.screener.engine import (
    load_config,
    apply_filters,
    run_preset
)
from src.screener.scoring import (
    build_scoring_dataset,
    load_scoring_data
)

# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
)

OUTPUT_PATH = (
    OUTPUT_DIR
    / "screener_output.xlsx"
)


# ============================================================
# SHEET NAMES
# ============================================================

SHEET_NAMES = {

    "quality_compounder":
        "Quality Compounder",

    "value_pick":
        "Value Pick",

    "growth_accelerator":
        "Growth Accelerator",

    "dividend_champion":
        "Dividend Champion",

    "debt_free_blue_chip":
        "Debt Free Blue Chip",

    "turnaround_watch":
        "Turnaround Watch"
}


# ============================================================
# EXPORT COLUMNS
# ============================================================

EXPORT_COLUMNS = [

    "company_id",
    "year",
    "broad_sector",

    # Profitability
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",

    # Leverage and efficiency
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",

    # Cash quality
    "free_cash_flow_cr",
    "fcf_cagr_5yr",
    "cfo_pat_ratio_5yr",

    # Growth
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    # Dividend
    "dividend_payout_ratio_pct",
    "dividend_yield_pct",

    # Valuation
    "pe_ratio",
    "pb_ratio",
    "market_cap_crore",

    # Scale
    "sales",
    "net_profit",

    # Scores
    "composite_quality_score_v2",
    "sector_relative_score"
]
# ============================================================
# COLOURS
# ============================================================

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7"
)


# ============================================================
# BUILD PRESET DATA
# ============================================================

def apply_de_declining_filter(
    result
):

    # ========================================================
    # LOAD FULL HISTORICAL DATA
    # ========================================================

    history = load_scoring_data()


    # ========================================================
    # REMOVE TTM
    # ========================================================

    history = history[
        history["year"] != "TTM"
    ].copy()


    # ========================================================
    # EXTRACT YEAR AND MONTH
    # ========================================================

    history["fiscal_year"] = (
        history["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0]
    )


    history["fiscal_year"] = pd.to_numeric(
        history["fiscal_year"],
        errors="coerce"
    )


    month_map = {

        "Mar": 3,

        "Jun": 6,

        "Sep": 9,

        "Dec": 12
    }


    history["fiscal_month"] = (
        history["year"]
        .astype(str)
        .str[:3]
        .map(month_map)
        .fillna(0)
    )


    # ========================================================
    # SORT CHRONOLOGICALLY
    # ========================================================

    history = history.sort_values(
        by=[
            "company_id",
            "fiscal_year",
            "fiscal_month"
        ]
    )


    # ========================================================
    # PREVIOUS PERIOD D/E
    # ========================================================

    history["previous_de"] = (
        history
        .groupby("company_id")[
            "debt_to_equity"
        ]
        .shift(1)
    )


    # ========================================================
    # DECLINING D/E FLAG
    # ========================================================

    history["de_declining"] = (

        history["debt_to_equity"].notna()

        &

        history["previous_de"].notna()

        &

        (
            history["debt_to_equity"]
            <
            history["previous_de"]
        )
    )


    # ========================================================
    # GET LATEST ANNUAL STATUS
    # ========================================================

    latest_status = (

        history

        .groupby(
            "company_id",
            as_index=False
        )

        .tail(1)

        [
            [
                "company_id",
                "previous_de",
                "debt_to_equity",
                "de_declining"
            ]
        ]

        .copy()
    )


    # ========================================================
    # KEEP DECLINING COMPANIES
    # ========================================================

    declining_companies = latest_status[
        latest_status["de_declining"]
    ][
        "company_id"
    ]


    filtered = result[
        result["company_id"].isin(
            declining_companies
        )
    ].copy()


    return filtered


def build_preset_results():

    scoring_df = build_scoring_dataset()

    preset_results = {}

    config = load_config()

    for preset_name in SHEET_NAMES:

        filters = config[
            "presets"
        ][
            preset_name
        ]

        # ====================================================
        # TURNAROUND WATCH
        # ====================================================

        if preset_name == "turnaround_watch":

            # Apply only standard threshold filters here.
            # D/E declining requires historical comparison
            # and is handled separately.

            standard_filters = {

                filter_name: threshold

                for filter_name, threshold in filters.items()

                if filter_name != "de_declining"
            }

            result = apply_filters(
                df=scoring_df,
                filters=standard_filters,
                config=config
            )

            # ------------------------------------------------
            # Historical D/E declining check
            # ------------------------------------------------

            result = apply_de_declining_filter(
                result
            )

        else:

            result = apply_filters(
                df=scoring_df,
                filters=filters,
                config=config
            )


        result = result.sort_values(
            by="composite_quality_score_v2",
            ascending=False,
            na_position="last"
        )


        available_columns = [

            column

            for column in EXPORT_COLUMNS

            if column in result.columns
        ]


        preset_results[
            preset_name
        ] = result[
            available_columns
        ].copy()


    return preset_results
# ============================================================
# WRITE WORKBOOK
# ============================================================

def write_workbook(
    preset_results
):

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True
    )


    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl"
    ) as writer:


        for preset_name, df in preset_results.items():

            sheet_name = SHEET_NAMES[
                preset_name
            ]


            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )


# ============================================================
# CHECK FILTER CONDITION
# ============================================================

def condition_passes(
    value,
    operator,
    threshold
):

    if pd.isna(value):

        return False


    if operator == "min":

        return value >= threshold


    if operator == "max":

        return value <= threshold


    if operator == "equal":

        return value == threshold


    return False


# ============================================================
# FORMAT WORKBOOK
# ============================================================

def format_workbook():

    config = load_config()


    workbook = load_workbook(
        OUTPUT_PATH
    )


    for preset_name, sheet_name in SHEET_NAMES.items():

        worksheet = workbook[
            sheet_name
        ]


        # ----------------------------------------------------
        # Header formatting
        # ----------------------------------------------------

        for cell in worksheet[1]:

            cell.fill = HEADER_FILL

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )


        # ----------------------------------------------------
        # Freeze header
        # ----------------------------------------------------

        worksheet.freeze_panes = "A2"


        # ----------------------------------------------------
        # Auto filter
        # ----------------------------------------------------

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


        # ----------------------------------------------------
        # Column map
        # ----------------------------------------------------

        column_map = {

            cell.value: cell.column

            for cell in worksheet[1]
        }


        filters = config[
            "presets"
        ][
            preset_name
        ]


        metric_config = config[
            "metrics"
        ]


        # ----------------------------------------------------
        # Threshold formatting
        # ----------------------------------------------------

        for filter_name, threshold in filters.items():

            if filter_name not in metric_config:

                continue


            rule = metric_config[
                filter_name
            ]


            column_name = rule[
                "column"
            ]


            operator = rule[
                "operator"
            ]


            if column_name not in column_map:

                continue


            excel_column = column_map[
                column_name
            ]


            for row_number in range(
                2,
                worksheet.max_row + 1
            ):


                cell = worksheet.cell(
                    row=row_number,
                    column=excel_column
                )


                value = cell.value


                passed = condition_passes(
                    value,
                    operator,
                    threshold
                )


                if passed:

                    cell.fill = GREEN_FILL

                else:

                    cell.fill = RED_FILL


        # ----------------------------------------------------
        # Column widths
        # ----------------------------------------------------

        for column_cells in worksheet.columns:

            max_length = 0


            column_letter = (
                column_cells[0]
                .column_letter
            )


            for cell in column_cells:

                value = cell.value


                if value is None:

                    continue


                max_length = max(
                    max_length,
                    len(str(value))
                )


            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                28
            )


    workbook.save(
        OUTPUT_PATH
    )


# ============================================================
# MAIN EXPORT
# ============================================================

def export_screener_workbook():

    print(
        "\nBuilding preset results..."
    )


    preset_results = build_preset_results()


    for preset_name, result in preset_results.items():

        print(
            f"{preset_name:<25} "
            f"{len(result)} companies"
        )


    print(
        "\nWriting Excel workbook..."
    )


    write_workbook(
        preset_results
    )


    print(
        "Applying workbook formatting..."
    )


    format_workbook()


    print(
        "\nScreener workbook generated:"
    )


    print(
        OUTPUT_PATH
    )


# ============================================================
# DIRECT RUN
# ============================================================

if __name__ == "__main__":

    export_screener_workbook()