from pathlib import Path
import re
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = (
    PROJECT_ROOT
    / "data"
    / "db"
    / "nifty100.db"
)

OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
)

OUTPUT_PATH = (
    OUTPUT_DIR
    / "peer_comparison.xlsx"
)


# ============================================================
# METRIC CONFIGURATION
# ============================================================

METRIC_COLUMNS = [

    "net_profit_margin_pct",

    "operating_profit_margin_pct",

    "return_on_equity_pct",

    "return_on_capital_employed_pct",

    "debt_to_equity",

    "interest_coverage",

    "asset_turnover",

    "free_cash_flow_cr",

    "capex_cr",

    "earnings_per_share",

    "book_value_per_share",

    "dividend_payout_ratio_pct",

    "total_debt_cr",

    "cash_from_operations_cr",

    "revenue_cagr_3yr",

    "revenue_cagr_5yr",

    "pat_cagr_5yr",

    "eps_cagr_5yr",

    "market_cap_crore",

    "dividend_yield_pct"
]


# ============================================================
# PERCENTILE METRIC MAPPING
# ============================================================

PERCENTILE_METRICS = {

    "roe":
        "roe_percentile",

    "roce":
        "roce_percentile",

    "net_profit_margin":
        "net_profit_margin_percentile",

    "debt_to_equity":
        "debt_to_equity_percentile",

    "free_cash_flow":
        "free_cash_flow_percentile",

    "pat_cagr_5yr":
        "pat_cagr_5yr_percentile",

    "revenue_cagr_5yr":
        "revenue_cagr_5yr_percentile",

    "eps_cagr_5yr":
        "eps_cagr_5yr_percentile",

    "interest_coverage":
        "interest_coverage_percentile",

    "asset_turnover":
        "asset_turnover_percentile"
}


# ============================================================
# COLOURS
# ============================================================

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C"
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966"
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7"
)

SUMMARY_FILL = PatternFill(
    fill_type="solid",
    fgColor="E7E6E6"
)


THIN_BORDER = Border(

    left=Side(
        style="thin"
    ),

    right=Side(
        style="thin"
    ),

    top=Side(
        style="thin"
    ),

    bottom=Side(
        style="thin"
    )
)


# ============================================================
# LOAD PEER GROUPS
# ============================================================

def load_peer_groups(
    conn
):

    return pd.read_sql(

        """
        SELECT
            peer_group_name,
            company_id,
            is_benchmark
        FROM peer_groups
        ORDER BY
            peer_group_name,
            company_id
        """,

        conn
    )


# ============================================================
# LOAD COMPANY NAMES
# ============================================================

def load_companies(
    conn
):

    return pd.read_sql(

        """
        SELECT
            company_id,
            company_name
        FROM companies
        """,

        conn
    )


# ============================================================
# LOAD FINANCIAL RATIOS
# ============================================================

def load_financial_ratios(
    conn
):

    return pd.read_sql(

        """
        SELECT *
        FROM financial_ratios
        """,

        conn
    )


# ============================================================
# LOAD MARKET DATA
# ============================================================

def load_market_data(
    conn
):

    return pd.read_sql(

        """
        SELECT
            company_id,
            year,
            market_cap_crore,
            enterprise_value_crore,
            pe_ratio,
            pb_ratio,
            ev_ebitda,
            dividend_yield_pct
        FROM market_cap
        """,

        conn
    )


# ============================================================
# LOAD PEER PERCENTILES
# ============================================================

def load_peer_percentiles(
    conn
):

    return pd.read_sql(

        """
        SELECT
            company_id,
            peer_group_name,
            metric,
            value,
            percentile_rank,
            year
        FROM peer_percentiles
        """,

        conn
    )


# ============================================================
# FISCAL YEAR EXTRACTION
# ============================================================

def extract_fiscal_year(
    value
):

    if pd.isna(value):

        return None


    match = re.search(
        r"(\d{4})",
        str(value)
    )


    if match:

        return int(
            match.group(1)
        )


    return None


# ============================================================
# LATEST ANNUAL FINANCIAL RATIOS
# ============================================================

def get_latest_annual_ratios(
    ratios
):

    df = ratios[
        ratios["year"] != "TTM"
    ].copy()


    df["fiscal_year"] = (
        df["year"]
        .apply(
            extract_fiscal_year
        )
    )


    month_map = {

        "Mar": 3,

        "Jun": 6,

        "Sep": 9,

        "Dec": 12
    }


    df["fiscal_month"] = (

        df["year"]
        .astype(str)
        .str[:3]
        .map(month_map)
        .fillna(0)
    )


    df = df.sort_values(

        by=[

            "company_id",

            "fiscal_year",

            "fiscal_month"
        ]
    )


    latest = (

        df
        .groupby(
            "company_id",
            as_index=False
        )
        .tail(1)
        .reset_index(
            drop=True
        )
    )


    latest = latest.drop(

        columns=[

            "fiscal_year",

            "fiscal_month"
        ],

        errors="ignore"
    )


    return latest


# ============================================================
# LATEST MARKET DATA
# ============================================================

def get_latest_market_data(
    market_data
):

    df = market_data.copy()


    df["market_year"] = pd.to_numeric(

        df["year"],

        errors="coerce"
    )


    df = df.sort_values(

        by=[

            "company_id",

            "market_year"
        ]
    )


    latest = (

        df
        .groupby(
            "company_id",
            as_index=False
        )
        .tail(1)
        .reset_index(
            drop=True
        )
    )


    latest = latest.drop(

        columns=[

            "year",

            "market_year"
        ],

        errors="ignore"
    )


    return latest


# ============================================================
# BUILD PERCENTILE WIDE TABLE
# ============================================================

def build_percentile_wide(
    percentiles
):

    wide = (

        percentiles
        .pivot_table(

            index=[

                "company_id",

                "peer_group_name"
            ],

            columns="metric",

            values="percentile_rank",

            aggfunc="first"
        )
        .reset_index()
    )


    wide.columns.name = None


    rename_map = {

        metric:
            percentile_column

        for metric, percentile_column
        in PERCENTILE_METRICS.items()
    }


    wide = wide.rename(
        columns=rename_map
    )


    return wide


# ============================================================
# BUILD REPORT DATASET
# ============================================================

def build_report_dataset():

    conn = sqlite3.connect(
        DB_PATH
    )


    peer_groups = load_peer_groups(
        conn
    )

    companies = load_companies(
        conn
    )

    ratios = load_financial_ratios(
        conn
    )

    market_data = load_market_data(
        conn
    )

    percentiles = load_peer_percentiles(
        conn
    )


    conn.close()


    # --------------------------------------------------------
    # Latest annual ratio records
    # --------------------------------------------------------

    latest_ratios = (
        get_latest_annual_ratios(
            ratios
        )
    )


    # --------------------------------------------------------
    # Latest market records
    # --------------------------------------------------------

    latest_market = (
        get_latest_market_data(
            market_data
        )
    )


    # --------------------------------------------------------
    # Percentile wide format
    # --------------------------------------------------------

    percentile_wide = (
        build_percentile_wide(
            percentiles
        )
    )


    # --------------------------------------------------------
    # Main merge
    # --------------------------------------------------------

    df = peer_groups.merge(

        companies,

        on="company_id",

        how="left"
    )


    df = df.merge(

        latest_ratios,

        on="company_id",

        how="left"
    )


    df = df.merge(

        latest_market,

        on="company_id",

        how="left"
    )


    df = df.merge(

        percentile_wide,

        on=[

            "company_id",

            "peer_group_name"
        ],

        how="left"
    )


    return df


# ============================================================
# REPORT COLUMN ORDER
# ============================================================

def get_report_columns(
    df
):

    base_columns = [

        "company_id",

        "company_name",

        "year",

        "is_benchmark"
    ]


    percentile_columns = list(
        PERCENTILE_METRICS.values()
    )


    desired_columns = (

        base_columns

        +

        METRIC_COLUMNS

        +

        percentile_columns
    )


    available_columns = [

        column

        for column in desired_columns

        if column in df.columns
    ]


    return available_columns


# ============================================================
# ADD MEDIAN SUMMARY ROW
# ============================================================

def add_summary_row(
    group_df
):

    df = group_df.copy()


    summary = {

        column: None

        for column in df.columns
    }


    summary[
        "company_id"
    ] = "MEDIAN"


    summary[
        "company_name"
    ] = "Peer Group Median"


    summary[
        "year"
    ] = ""


    summary[
        "is_benchmark"
    ] = ""


    numeric_columns = (

        df
        .select_dtypes(
            include="number"
        )
        .columns
        .tolist()
    )


    for column in numeric_columns:

        if column == "is_benchmark":

            continue


        summary[
            column
        ] = df[
            column
        ].median(
            skipna=True
        )


    summary_df = pd.DataFrame(
        [summary]
    )


    return pd.concat(

        [
            df,
            summary_df
        ],

        ignore_index=True
    )


# ============================================================
# SAFE EXCEL SHEET NAME
# ============================================================

def safe_sheet_name(
    name
):

    invalid_characters = [

        "\\",

        "/",

        "*",

        "?",

        ":",

        "[",

        "]"
    ]


    result = str(
        name
    )


    for character in invalid_characters:

        result = result.replace(
            character,
            "-"
        )


    return result[:31]


# ============================================================
# BUILD GROUP REPORTS
# ============================================================

def build_group_reports():

    dataset = build_report_dataset()


    report_columns = (
        get_report_columns(
            dataset
        )
    )


    reports = {}


    for peer_group_name, group_df in dataset.groupby(
        "peer_group_name",
        sort=True
    ):


        report_df = group_df[
            report_columns
        ].copy()


        report_df = report_df.sort_values(

            by=[

                "is_benchmark",

                "company_id"
            ],

            ascending=[

                False,

                True
            ]
        )


        report_df = add_summary_row(
            report_df
        )


        reports[
            peer_group_name
        ] = report_df


    return reports


# ============================================================
# WRITE RAW WORKBOOK
# ============================================================

def write_workbook(
    reports
):

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True
    )


    with pd.ExcelWriter(

        OUTPUT_PATH,

        engine="openpyxl"

    ) as writer:


        for peer_group_name, report_df in reports.items():


            sheet_name = safe_sheet_name(
                peer_group_name
            )


            report_df.to_excel(

                writer,

                sheet_name=sheet_name,

                index=False
            )


# ============================================================
# PERCENTILE COLOUR
# ============================================================

def get_percentile_fill(
    value
):

    if value is None:

        return None


    try:

        value = float(
            value
        )

    except (
        TypeError,
        ValueError
    ):

        return None


    if value >= 0.75:

        return GREEN_FILL


    if value <= 0.25:

        return RED_FILL


    return YELLOW_FILL


# ============================================================
# FORMAT WORKBOOK
# ============================================================

def format_workbook():

    workbook = load_workbook(
        OUTPUT_PATH
    )


    percentile_columns = set(
        PERCENTILE_METRICS.values()
    )


    for worksheet in workbook.worksheets:


        # ----------------------------------------------------
        # Freeze header
        # ----------------------------------------------------

        worksheet.freeze_panes = "A2"


        # ----------------------------------------------------
        # Auto filter
        # ----------------------------------------------------

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


        # ----------------------------------------------------
        # Header formatting
        # ----------------------------------------------------

        for cell in worksheet[1]:

            cell.fill = HEADER_FILL

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(

                horizontal="center",

                vertical="center",

                wrap_text=True
            )

            cell.border = THIN_BORDER


        # ----------------------------------------------------
        # Column map
        # ----------------------------------------------------

        column_map = {

            cell.value:
                cell.column

            for cell in worksheet[1]
        }


        # ----------------------------------------------------
        # Identify summary row
        # ----------------------------------------------------

        summary_row = (
            worksheet.max_row
        )


        # ----------------------------------------------------
        # Percentile colour coding
        # ----------------------------------------------------

        for percentile_column in percentile_columns:


            if percentile_column not in column_map:

                continue


            column_number = column_map[
                percentile_column
            ]


            for row_number in range(

                2,

                summary_row

            ):


                cell = worksheet.cell(

                    row=row_number,

                    column=column_number
                )


                fill = get_percentile_fill(
                    cell.value
                )


                if fill is not None:

                    cell.fill = fill


                cell.number_format = "0.00%"


        # ----------------------------------------------------
        # Benchmark row highlight
        # ----------------------------------------------------

        benchmark_column = column_map.get(
            "is_benchmark"
        )


        if benchmark_column is not None:


            for row_number in range(

                2,

                summary_row

            ):


                benchmark_value = worksheet.cell(

                    row=row_number,

                    column=benchmark_column

                ).value


                if benchmark_value == 1:


                    for cell in worksheet[
                        row_number
                    ]:

                        cell.fill = BENCHMARK_FILL

                        cell.font = Font(
                            bold=True
                        )


        # ----------------------------------------------------
        # Summary row formatting
        # ----------------------------------------------------

        for cell in worksheet[
            summary_row
        ]:

            cell.fill = SUMMARY_FILL

            cell.font = Font(
                bold=True
            )

            cell.border = THIN_BORDER


        # ----------------------------------------------------
        # Borders and alignment
        # ----------------------------------------------------

        for row in worksheet.iter_rows(

            min_row=2,

            max_row=worksheet.max_row

        ):

            for cell in row:

                cell.border = THIN_BORDER

                cell.alignment = Alignment(

                    vertical="center"
                )


        # ----------------------------------------------------
        # Numeric formatting
        # ----------------------------------------------------

        for header_name, column_number in column_map.items():


            if header_name in percentile_columns:

                continue


            if header_name in [

                "company_id",

                "company_name",

                "year",

                "is_benchmark"

            ]:

                continue


            for row_number in range(

                2,

                worksheet.max_row + 1

            ):

                worksheet.cell(

                    row=row_number,

                    column=column_number

                ).number_format = "0.00"


        # ----------------------------------------------------
        # Column widths
        # ----------------------------------------------------

        for column_cells in worksheet.columns:


            column_letter = (
                column_cells[0]
                .column_letter
            )


            max_length = 0


            for cell in column_cells:


                if cell.value is None:

                    continue


                max_length = max(

                    max_length,

                    len(
                        str(
                            cell.value
                        )
                    )
                )


            worksheet.column_dimensions[
                column_letter
            ].width = min(

                max_length + 2,

                24
            )


        # ----------------------------------------------------
        # Row height
        # ----------------------------------------------------

        worksheet.row_dimensions[
            1
        ].height = 35


    workbook.save(
        OUTPUT_PATH
    )


# ============================================================
# VALIDATE WORKBOOK
# ============================================================

def validate_workbook():

    workbook = load_workbook(
        OUTPUT_PATH,
        data_only=False
    )


    print(
        "\n"
        + "=" * 100
    )

    print(
        "DAY 20 — PEER COMPARISON WORKBOOK VALIDATION"
    )

    print(
        "=" * 100
    )


    print(
        "\nSheet Count:",
        len(
            workbook.sheetnames
        )
    )


    print(
        "\nSheets:"
    )


    for sheet_name in workbook.sheetnames:

        worksheet = workbook[
            sheet_name
        ]


        headers = [

            cell.value

            for cell in worksheet[1]
        ]


        company_id_column = (
            headers.index(
                "company_id"
            )
            + 1
        )


        benchmark_column = (
            headers.index(
                "is_benchmark"
            )
            + 1
        )


        company_count = 0

        benchmark_count = 0

        summary_found = False


        for row_number in range(

            2,

            worksheet.max_row + 1

        ):


            company_id = worksheet.cell(

                row=row_number,

                column=company_id_column

            ).value


            benchmark_value = worksheet.cell(

                row=row_number,

                column=benchmark_column

            ).value


            if company_id == "MEDIAN":

                summary_found = True

                continue


            if company_id is not None:

                company_count += 1


            if benchmark_value == 1:

                benchmark_count += 1


        print(

            f"{sheet_name:<25} "

            f"Companies={company_count:<3} "

            f"Benchmark={benchmark_count} "

            f"Median Row={summary_found}"
        )


    if len(
        workbook.sheetnames
    ) == 11:

        print(
            "\nSheet Count Validation: PASS"
        )

    else:

        print(
            "\nSheet Count Validation: FAIL"
        )


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "\n"
        + "=" * 100
    )

    print(
        "DAY 20 — PEER COMPARISON EXCEL REPORT"
    )

    print(
        "=" * 100
    )


    print(
        "\nBuilding peer group reports..."
    )


    reports = build_group_reports()


    for peer_group_name, report_df in reports.items():

        company_count = (

            report_df[
                "company_id"
            ]
            !=
            "MEDIAN"

        ).sum()


        print(

            f"{peer_group_name:<25} "

            f"{company_count} companies"
        )


    print(
        "\nWriting workbook..."
    )


    write_workbook(
        reports
    )


    print(
        "Applying formatting..."
    )


    format_workbook()


    validate_workbook()


    print(
        "\nWorkbook generated:"
    )


    print(
        OUTPUT_PATH
    )


if __name__ == "__main__":

    main()